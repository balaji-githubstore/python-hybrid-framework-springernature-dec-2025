from typing import List, Dict

import openpyxl


class ExcelSource:

    @staticmethod
    def get_one_row(file_path, sheet_name, row_num: int) -> List[List]:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        return [[cell.value for cell in sheet[row_num]]]

    @staticmethod
    def get_exact_rows(file_path, sheet_name, row_num: List[int]) -> List[List]:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        final = []
        for row in row_num:
            final.append([cell.value for cell in sheet[row]])
        return final

    @staticmethod
    def get_all_rows_as_dicts(file_path, sheet_name) -> List[Dict]:
        """
        Return the sheet data as a list of dictionaries.
        Header = first row.
        """
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        rows = list(sheet.rows)
        header = [cell.value for cell in rows[0]]
        data_rows = rows[1:]

        result = []
        for row in data_rows:
            values = [cell.value for cell in row]
            result.append(dict(zip(header, values)))

        return result

